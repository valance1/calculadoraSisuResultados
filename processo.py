import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment

df = pd.read_excel('template.xlsx', header=None)

Notas_dict = {}
ids = df.iloc[0, 1:]  # Assuming IDs are in the first row, starting from the second column
Notas = df.iloc[2:, 1:]  # Assuming Notas start from the third row, starting from the second column
vagas = df.iloc[1, 1:]

# Replace NaN with empty string
Notas = Notas.fillna('')

for i, id in enumerate(ids):
    Notas_dict[id] = Notas.iloc[:, i].tolist()

sorted_Notas = []

for id, Notas_list in Notas_dict.items():
    for Nota in Notas_list:
        if Nota != '':  # Skip empty Notas
            sorted_Notas.append((float(Nota), id))

sorted_Notas.sort(reverse=True, key=lambda x: x[0])

#for Nota, id in sorted_Notas:
#    print(f"ID: {id}, Nota: {Nota}")

vacancies = int(vagas.iloc[0])
selected_students = sorted_Notas[:vacancies]
remaining_students = sorted_Notas[vacancies:]

print("\nSelected Students FOR AC:")
for Nota, id in selected_students:
    print(f"ID: {id}, Nota: {Nota}")

# print("\nRemaining Students:")
grouped_remaining_students = {}

for i, id in enumerate(ids):
    grouped_remaining_students[id] = []

for Nota, id in remaining_students:
    grouped_remaining_students[id].append(Nota)

#print("\nGrouped Remaining Students:")
#for id, Notas_list in grouped_remaining_students.items():
#    print(f"ID: {id}, Notas: {Notas_list}")

# Sort each group and respect the vacancy limit, except for 'AC'
chamada_regular = {}
for i, id in enumerate(ids):
    if id == 'AC':
        sorted_group = sorted(grouped_remaining_students[id], reverse=True)
    else:
        limit = int(vagas.iloc[i])
        sorted_group = sorted(grouped_remaining_students[id], reverse=True)[:limit]
    chamada_regular[id] = sorted_group

# Remove selected students from remaining, except for 'AC'
for id, Notas_list in chamada_regular.items():
    if id == 'AC':
        continue
    for Nota in Notas_list:
        if Nota in grouped_remaining_students[id]:
            grouped_remaining_students[id].remove(Nota)

print("\nFinal Groups:")
for id, Notas_list in chamada_regular.items():
    print(f"ID: {id}, Notas: {Notas_list}")

print("\nUpdated Grouped Remaining Students:")
for id, Notas_list in grouped_remaining_students.items():
    print(f"ID: {id}, Notas: {Notas_list}")

# Create a waiting list from the remaining students
lista_de_espera = []

for id, Notas_list in grouped_remaining_students.items():
    for Nota in Notas_list:
        lista_de_espera.append((float(Nota), id))

lista_de_espera.sort(reverse=True, key=lambda x: x[0])

print("\nWaiting List:")
for Nota, id in lista_de_espera:
    print(f"ID: {id}, Nota: {Nota}")


# Create a DataFrame for selected students
selected_df = pd.DataFrame(selected_students, columns=['Nota', 'ID'])

# Create a DataFrame for final groups
chamada_regular_df = pd.DataFrame.from_dict(chamada_regular, orient='index').transpose()
chamada_regular_df.insert(0, 'Notas', selected_df['Nota'])
chamada_regular_df.drop(chamada_regular_df.columns[1], axis=1, inplace=True)

# Create a DataFrame for waiting list
lista_de_espera_df = pd.DataFrame(lista_de_espera, columns=['Nota', 'ID'])

wb = openpyxl.load_workbook('results.xlsx')
chamada_regular_sheet = wb['Chamada Regular']
lista_de_espera_sheet = wb['Lista de Espera']

# Add titles
chamada_regular_sheet.insert_rows(1)
chamada_regular_sheet['A1'] = 'Chamada Regular'
chamada_regular_sheet['A1'].font = Font(size=14, bold=True)
chamada_regular_sheet['A1'].alignment = Alignment(horizontal='center')
chamada_regular_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=chamada_regular_sheet.max_column)

lista_de_espera_sheet.insert_rows(1)
lista_de_espera_sheet['A1'] = 'Lista de Espera'
lista_de_espera_sheet['A1'].font = Font(size=14, bold=True)
lista_de_espera_sheet['A1'].alignment = Alignment(horizontal='center')
lista_de_espera_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=lista_de_espera_sheet.max_column)

# Adjust column widths
for sheet in [chamada_regular_sheet, lista_de_espera_sheet]:
    for col in sheet.columns:
        max_length = 0
        if isinstance(col[0], openpyxl.cell.cell.MergedCell):
            continue
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

# Save the workbook
wb.save('results.xlsx')
